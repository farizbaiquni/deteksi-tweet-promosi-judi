import pandas as pd
from flask import Flask, render_template, request, jsonify
from flask_cors import CORS
import joblib
from nltk.corpus import stopwords
import re
from werkzeug.utils import secure_filename
import io
import openpyxl
from io import BytesIO

app = Flask(__name__)

CORS(app)

CORS(app, origins="http://localhost:3000", allow_headers=["Content-Type", "Authorization"])

# ================== PRE-PROCESSING ==================
stop_words = set(stopwords.words('indonesian'))

#text cleaning function
def pre_processing(text):
    
    text = text.lower()
    
    result = ""
    
    #removing any special character
    text=re.sub(r"http.+|https.+"," ",text)
    text=re.sub(r"[-()\"#!@$%^&*{}?.,:]", " ",text)
    text=re.sub(r"\s+"," ",text)
    text=re.sub(r"[^A-Za-z0-9]+", " ", text)
    text=re.sub(r"\d+", " ", text)
    text=re.sub(r"\b\w\b", " ", text)
    text=re.sub(r"\s+", " ", text)
    
    for word in text.split():
        if word not in stop_words:
            result+=word + " "
    
    return result

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls', 'csv'}

# Memuat model dari file
loaded_model = joblib.load("model_prediksi_tweet_promosi_judi.joblib")

@app.route("/check")
def check():
    result = {"result": "OK"}
    return jsonify(result), 200


@app.route("/predict", methods=["POST"])
def predict():
    if request.method == "POST":
        text = request.json["text"]
        text = pre_processing(text)
        predicted_category = loaded_model.predict([text])[0]
        result = {"prediction": int(predicted_category)}
        return jsonify(result), 200
    
@app.route("/analisis", methods=["POST"])
def analisis():
    if request.method == "POST":
        # Memeriksa apakah ada file yang dikirim dengan permintaan POST
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
    
        file = request.files['file']
        
        if file and allowed_file(file.filename):
            
            if file.filename.rsplit('.', 1)[1].lower() == 'csv':
                # Membaca file CSV dari BytesIO
                csv_content = file.read()
                df = pd.read_csv(BytesIO(csv_content))

                # Menyimpan DataFrame ke dalam file Excel
                excel_content = BytesIO()
                df.to_excel(excel_content, index=False)
                excel_content.seek(0)
                
                # Membaca workbook dari BytesIO
                wb = openpyxl.load_workbook(excel_content)  
            else:
                file_content = file.read()
                wb = openpyxl.load_workbook(io.BytesIO(file_content))
            
            # Mendapatkan nama sheet pertama
            sheet_name = wb.sheetnames[0]
            sheet = wb[sheet_name]
    
            # Mendapatkan header (kolom pertama)
            header = [cell.value for cell in sheet[1]]
    
            # Menambahkan kolom "prediction" ke header
            header.append("prediction")
    
            # Mengonversi setiap baris menjadi dictionary dengan tambahan kolom "prediction"
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                prediction_value = row[2] 
                prediction_value = pre_processing(prediction_value)
                aaa = loaded_model.predict([prediction_value])[0]
                prediction_value = int(aaa)
                row_data = list(row) + [prediction_value]  
                data.append(dict(zip(header, row_data)))
                
            # Membalik urutan data    
            data = data[::-1]
    
            # Mengembalikan JSON
            return jsonify({'result': data}), 200
        
        else:
            return jsonify({'error': 'File tidak diizinkan'}), 400
        


if __name__ == "__main__":
    app.run(debug=True)
